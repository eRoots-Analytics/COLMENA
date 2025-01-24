import openpyxl
import numpy as np
from GridCalEngine import Bus
import matplotlib.pyplot as plt
import os, sys
import control as ctrl
current_directory = os.path.dirname(os.path.abspath(__file__))
two_levels_up = os.path.dirname(os.path.dirname(current_directory))
sys.path.insert(0, two_levels_up)
import andes as ad

def build_new_system_legacy(system, new_model_name = 'REDUAL'):
    system_to = ad.System()
    system_dict = system.as_dict()
    gen_model = 'GENROU'
    gen_dependencies = ['IEEEST', 'TGOV1N', 'IEEEX1']
    n_dual = 1
    n_genrou = system.GENROU.n - n_dual
    
    
    for model, param_dict in system_dict.items():
        #if n_genrou is 0 we just change all generators for 
        if model == gen_model and n_genrou == 0:
            model = new_model_name
        elif model in gen_dependencies and n_genrou == 0:
            continue
        
        elif model == gen_model and n_genrou > 0:
            _ = 0
            
        elif model in gen_dependencies and n_genrou > 0:
            for i in range(n_genrou):
                new_dict = {key: value[i] for key, value in param_dict.items()}
                system_to.add(model, new_dict)
            continue
        
        for i in range(len(param_dict['u'])):        
            new_dict = {key: value[i] for key, value in param_dict.items() if isinstance(value, list) or isinstance(value, np.ndarray)}
            new_dict_ = 0
            generator_like = ['GENROU', 'REGCV1', 'REGCA1']
            if i < n_genrou and model in generator_like:
                model = 'GENROU'
            elif i < n_genrou + n_dual and model in generator_like:
                model = 'REDUAL'
            system_to.add(model, new_dict)        
            
    return system_to

def build_new_system(system, model_swap = {'REDUAL':['REGCV1','REGCP1']}):
    system_to = ad.System()
    system_dict = system.as_dict()
    
    for model, param_dict in system_dict.items():
        if model in model_swap.keys():
            for model_to in model_swap[model]:
                for i in range(len(param_dict['u'])):        
                    new_dict = {key: value[i] for key, value in param_dict.items() if isinstance(value, list) or isinstance(value, np.ndarray)}
                    system_to.add(model_to, new_dict)
        else:
            for i in range(len(param_dict['u'])):        
                new_dict = {key: value[i] for key, value in param_dict.items() if isinstance(value, list) or isinstance(value, np.ndarray)}
                system_to.add(model, new_dict)

    return system_to

def replace_in_file(file_path, output_path=None):
    """
    Replaces every occurrence of 'STAB2A' with 'IEEEST' in a file.

    Parameters:
    - file_path (str): Path to the input file.
    - output_path (str): Path to the output file. If None, the input file is overwritten.

    Returns:
    - None
    """
    try:
        # Open the file and read its content
        with open(file_path, 'r') as file:
            content = file.read()

        # Replace occurrences of 'STAB2A' with 'IEEEST'
        #content = content.replace('STAB2A', 'ST2CUT')
        content = content.replace('IEEET2', 'IEEET1')

        # Determine the output file path
        if output_path is None:
            return
            output_path = file_path

        # Write the updated content back to the file
        with open(output_path, 'w') as file:
            file.write(content)

        print(f"Successfully replaced models in {output_path}")
    except FileNotFoundError:
        print(f"Error: File not found at {file_path}")
    except Exception as e:
        print(f"An error occurred: {e}")

def add_sheet(workbook, sheet_name, data):

    # Create a new sheet
    if sheet_name in workbook.sheetnames:
        print(f'Sheet "{sheet_name}" already exists. It will be overwritten.')
        workbook.remove(workbook[sheet_name])
    new_sheet = workbook.create_sheet(title=sheet_name)

    # Write the header row
    for col_idx, key in enumerate(data.keys(), 1):
        new_sheet.cell(row=1, column=col_idx, value=key)

    # Write the data rows
    for row_idx in range(len(next(iter(data.values())))):  # Assuming all columns have the same length
        for col_idx, key in enumerate(data.keys(), 1):
            new_sheet.cell(row=row_idx + 2, column=col_idx, value=data[key][row_idx])

    return workbook

def set_config(system, setup=1, device_uid=0, toggle_model= None):
    tds_object = system.TDS_colmena
    def f_condition(mdl, device_idx, colmena_device =[1], epsilon = 0.003, dae_t = 0):
        uid = mdl.idx2uid(device_idx)
        if device_idx not in colmena_device:
            return False
        if mdl.omega.v[uid] > 1-epsilon and mdl.omega.v[uid]<1+epsilon:
            return False
        return True

    def f_condition2(mdl, device_idx, colmena_device =[1], epsilon = 0.003, dae_t=0):
        uid = mdl.idx2uid(device_idx)
        if device_idx not in colmena_device:
            return False
        if dae_t < 10:
            return False
        return True
    
    
    if getattr(tds_object.config, 'changes', None) is None:
        tds_object.config.changes = {}
    if setup == 1:
        tds_object.config.instructions = (f_condition, lambda x: 0*x + 13, lambda x: 0*x + 130)
        tds_object.config.instruction_msg = {}
        tds_object.config.instruction_msg["GENROU"] = (f_condition, lambda x: 0*x + 13, lambda x: 0*x + 130)
        tds_object.config.param = 'M'
        tds_object.config.device_uid = device_uid
        tds_object.config.changes[setup] = tds_object.config.instructions
    if setup == 2:
        tds_object.config.instructions = (f_condition2, lambda x: 0.1, lambda x: 0.2)
        tds_object.config.instruction_msg = {}
        tds_object.config.instruction_msg["PVD1"] = (f_condition2, lambda x: 0.1, lambda x: 0.2)
        tds_object.config.param = 'gammap'
        tds_object.config.device_uid = device_uid
        tds_object.config.changes[setup] = tds_object.config.instructions
    if setup == 3:
        tds_object.config.instructions = (f_condition2, lambda x: 1, lambda x: 0)
        tds_object.config.instruction_msg = {}
        tds_object.config.instruction_msg["Toggle_Line"] = (f_condition2, lambda x: 1, lambda x: 0)
        tds_object.config.param = 'connect'
        tds_object.config.device_uid = device_uid
        tds_object.config.changes[setup] = tds_object.config.instructions
    if setup == 4:
        tds_object.config.instructions = (f_condition2, lambda x: 1, lambda x: 0)
        tds_object.config.instruction_msg = {}
        tds_object.config.instruction_msg["GENROU_2"] = (f_condition2, lambda x: 1, lambda x: 0)
        tds_object.config.param = 'u'
        tds_object.config.device_uid = device_uid
        tds_object.config.changes[setup] = tds_object.config.instructions
    if setup == 5:
        tds_object.config.instructions = (f_condition2, lambda x: 0, lambda x: 999)
        tds_object.config.instruction_msg = {}
        tds_object.config.instruction_msg["Line"] = (f_condition2, lambda x: 0, lambda x: 999)
        tds_object.config.param = 'r'
        tds_object.config.device_uid = device_uid
        tds_object.config.changes[setup] = tds_object.config.instructions
    if setup == 6:
        tds_object.config.instructions = (f_condition2, lambda x: 1, lambda x: 0)
        tds_object.config.instruction_msg = {}
        tds_object.config.instruction_msg[toggle_model] = (f_condition2, lambda x: 1, lambda x: 0)
        tds_object.config.param = 'connect'
        tds_object.config.device_uid = device_uid
        tds_object.config.changes[setup] = tds_object.config.instructions
        
    if isinstance(setup, (list, np.ndarray)):
        for i, set in enumerate(setup):
            set_config(system, setup=set)
    return

def delete_last_row(workbook, sheet_name):
    sheet = workbook[sheet_name]
    
    # Get the maximum row number (last row)
    max_row = sheet.max_row
    max_column = sheet.max_column
    
    # Get the headers from the first row
    headers = [sheet.cell(row=1, column=i).value for i in range(1, max_column + 1)]
    
    # Get the data from the last row
    last_row_data = [sheet.cell(row=max_row, column=i).value for i in range(1, max_column + 1)]
    
    # Create the dictionary from headers and last row data
    last_row_dict = {headers[i]: [last_row_data[i]] for i in range(max_column)}
    
    # Delete the last row
    sheet.delete_rows(max_row)
    
    return last_row_dict

def transfer_grid_info(system_from, system_to):
    for model_from in system_from.models:
        model_from = getattr(system_from, model_from)
        if model_from.n == 0:
            continue
        model_name = model_from.class_name
        model_to = getattr(system_to, model_name)
        
        for var_name, var_from in model_from._states_and_ext().items():
            for i in range(model_from.n):
                var_to = getattr(model_to, var_name)
                var_to.v[i] = var_from.v[i]
                
        for var_name, var_from in model_from._algebs_and_ext().items():
            for i in range(model_from.n):
                var_to = getattr(model_to, var_name)
                var_to.v[i] = var_from.v[i]
        
        for var_name, var_from in model_from._all_params().items():
            for i in range(model_from.n):
                var_to = getattr(model_to, var_name)
                try:
                    var_to.v[i] = var_from.v[i]
                except:
                    _ = 0
        
        for var_name, var_from in model_from.discrete.items():
            for i in range(model_from.n):
                var_to = getattr(model_to, var_name)
                for flag in ['v', 'zu', 'zl', 'zi']:
                    try:
                        val = getattr(var_from, flag)[i]
                        getattr(var_to, flag)[i] = val
                    except:
                        _ = 0
    
    system_to.dae.t = system_from.dae.t
    return system_to

class PIcontroller:
    def __init__(self, **kwargs):
        self.params = {}
        self.dt = kwargs.get("dt", 0.01)
        self.Kp = kwargs.get("Kp", 0.1)  # Proportional gain
        self.Ki = kwargs.get("Ki", 0.1)  # Integral gain
        self.Lmin = kwargs.get("Lmin", -10.0)  # Minimum limit
        self.Lmax = kwargs.get("Lmax", 10.0)  # Maximum limit
        self.reference = kwargs.get("ref", 1.0)  
        self.is_delta = kwargs.get("ref", True)  
        self.model = kwargs.get("model", 'GENROU')  
        self.target_var = kwargs.get("target_var", 'paux0')  
        self.idx = kwargs.get("idx", 1)  

        # Transfer function for the PI controller: H(s) = Kp + Ki/s
        num = [self.Kp, self.Ki]
        den = [1, 0]  # s in the denominator
        self.pi_tf = ctrl.TransferFunction(num, den)

        # Initialize limiter function
        def limiter_min_max(input):
            if input < self.Lmin:
                res = self.Lmin
            elif input > self.Lmax:
                res = self.Lmax
            else:
                res = input
            return res
        self.limiter = limiter_min_max

         # Convert to state-space form
        self.pi_tf = ctrl.c2d(self.pi_tf, self.dt, method='zoh')
        self.pi_ss = ctrl.tf2ss(self.pi_tf)

        # Initialize states for the state-space representation
        self.x0 = np.zeros(self.pi_ss.A.shape[0])

    def apply(self, input, feedback = False):
        """
        Process a single input value through the PI controller with feedback.

        Args:
        - setpoint (float): The desired reference value.
        - measurement (float): The measured feedback value.

        Returns:
        - output_value (float): The controller's output.
        """
        # Calculate the error between the setpoint and the measurement
        if feedback:
            error = self.reference - input
        else:
            error = input

        # Simulate one step using the discrete-time transfer function
        self.pi_tf
        response= ctrl.forced_response(
            self.pi_tf,
            T = [0, self.dt],
            U=[error, error],  # Error signal as input
            X0 =self.x0  # Initial state
        )

        self.x0 = response.states[:, -1]
        y = response.outputs[-1]
    
        # Apply the limiter to the output
        res = self.limiter(y)
        return res

    def reset_state(self):
        self.x0 = 0*self.x0
        return
    
    def get_set_point(self, input, feedback = True):
        res = []
        output = self.apply(input, feedback=feedback)
        new_set_point = {}
        new_set_point['model'] = 'TGOV1'
        new_set_point['param'] = 'paux0'
        new_set_point['idx'] = self.idx
        new_set_point['value'] = output
        new_set_point['add'] = False

        if self.model != 'GENROU':
            new_set_point['model'] = self.model
            new_set_point['param'] = self.target_var
            new_set_point['value'] = output + self.reference*(1-self.is_delta)
        res.append(new_set_point)
        res.append(new_set_point)
        return res

class Stabilizer():
    def __init__(self, **kwargs):
        self.params = 0
        self.dt = 0.1
        self.idx = kwargs.get("idx", 1)
        self.T1 = kwargs.get("T1", 1.0)
        self.T2 = kwargs.get("T2", 1.0)
        self.T3 = kwargs.get("T3", 1.0)
        self.T4 = kwargs.get("T4", 1.0)
        self.T5 = kwargs.get("T5", 1.0)
        self.T6 = kwargs.get("T6", 1.0)
        self.A1 = kwargs.get("A1", 1.0)
        self.A2 = kwargs.get("A2", 1.0)
        self.A3 = kwargs.get("A3", 1.0)
        self.A4 = kwargs.get("A4", 1.0)
        self.A5 = kwargs.get("A5", 1.0)
        self.A6 = kwargs.get("A6", 1.0)
        self.Ks = kwargs.get("Ks", 1.0)
        self.Lmin = kwargs.get("Lmin", 1.0)
        self.Lmax = kwargs.get("Lmax", 1.0)
        self.Vlast = 0

        num_filter = [1, self.A1, self.A2]  # Numerator
        den_filter = np.polymul([1, self.A3, self.A4], [1, self.A5, self.A6])  # Denominator
        self.filter_tf = ctrl.TransferFunction(num_filter, den_filter)

        # Block 5: H(s) = (1 + s*T1) / (1 + s*T2)
        self.block5_tf = ctrl.TransferFunction([self.T1, 1], [self.T2, 1])

        # Block 6: H(s) = (1 + s*T3) / (1 + s*T4)
        self.block6_tf = ctrl.TransferFunction([self.T3, 1], [self.T4, 1])

        # Limiter dynamics (Block 7): H(s) = (Ks * s * T6) / (1 + s * T6)
        self.limiter_tf = ctrl.TransferFunction([self.Ks * self.T6, 0], [self.T6, 1])
        self.high_pass = ctrl.TransferFunction([self.Ks * self.T5, 0], [self.T5, 1])

        self.combined_tf = ctrl.series(self.filter_tf, self.high_pass, self.block5_tf, self.block6_tf, self.limiter_tf)
        # Combine all blocks in series

        self.combined_tf = ctrl.c2d(self.combined_tf, self.dt, method='zoh')


        def limiter_min_max(input):
            if input < self.Lmin:
                res = self.Lmin
            elif input > self.Lmax:
                res = self.Lmax
            else:
                res = input
            return res
        self.limiter = limiter_min_max

        # Initialize real-time simulation variables
        self.pi_ss = ctrl.tf2ss(self.combined_tf)

        # Initialize states for the state-space representation
        self.x0 = np.zeros(self.pi_ss.A.shape[0])
        #ctrl.bode(self.pi_ss, dB=True)
        #plt.show()

    #input is a single value
     # Process a single input value
    def apply(self, current_input):
        """
        Process a single input value through the controller.

        Args:
        - current_input (float): The input signal to the controller.

        Returns:
        - output_value (float): The controller's output.
        """
        # Simulate one step using the discrete-time transfer function
        """response = ctrl.forced_response(
            self.pi_ss, 
            T=[0, self.dt],  # Time window for this step
            U=[current_input, current_input],  # Input signal
            X0=self.x0  # Initial state
        )

        #We save the state and the output
        self.x0 = response.states[:, -1]
        y = response.outputs[-1]"""

        # Extract state-space matrices
        A, B, C, D = self.pi_ss.A, self.pi_ss.B, self.pi_ss.C, self.pi_ss.D

        # Compute next state and output for one timestep
        y = C @ self.x0 + D * current_input
        self.x0 = A @ self.x0 + B * current_input

        print('v stabilizer is', y[0,0])
        return self.limiter(y[0,0])
    
    def define_inputs_dict(self):
        vars = ['SW_s1', 'omega', 'SW_s3', 'tm0', 'Sn', 'Sb', 'SW_s4', 'tm', 'SW_s5', 'v']
        return vars

    def compute_input(self,**kwargs):
        vars = ['SW_s1', 'omega', 'SW_s3', 'tm0', 'Sn', 'Sb', 'SW_s4', 'tm', 'SW_s5', 'v']
        self.param = {}
        for var in vars:
            self.param[var] = kwargs.get(var, 1.0)
        
        sum1 = self.param['SW_s1']*(self.param['omega']-1)
        sum2 = self.param['SW_s3']*self.param['tm0']/(self.param['tm0'])
        sum3 = self.param['SW_s4']*(self.param['tm0'] - self.param['tm'])
        sum4 = self.param['SW_s5']*(self.param['v'])

        sig = sum1 + sum2 + sum3 + sum4
        return sig
    
    def get_set_point(self, input):
        res = []
        output = self.apply(input)
        new_set_point = {}
        new_set_point['model'] = 'EXDC2'
        new_set_point['param'] = 'v_aux'
        new_set_point['idx'] = self.idx
        new_set_point['value'] = self.Vlast
        new_set_point['add'] = False
        self.Vlast = output
        res.append(new_set_point)
        return res

class ActivePowerRegulator():
    def __init__(self, **kwargs):
        self.a = kwargs.get("a", 0.1)
        self.idx = kwargs.get("idx", 1.0)
        self.neighbors = kwargs.get("neighbors", 1.0)
        self.w0 = kwargs.get("ref", 1.0)

    def get_power(self, idx, system = None):
        if system is None:
            #when in colmena change this to access the data differently
            return
        
        uid = system.GENROU.idx2uid(idx)
        res = system.GENROU.Pe.v[uid]
        return res

    def compute_input(self, p_i, system =None):
        res = 0
        p_neighbors = [self.get_power(idx, system) for idx in self.neighbors]
        for p_j in p_neighbors:
            res += (p_j - p_i)
        
        return self.a*res
        
    def get_set_point(self, input, system):
        p_i = input
        d_w = self.compute_input(p_i, system)

        new_set_point = {}
        new_set_point['model'] = 'TGOV1'
        new_set_point['param'] = 'wref0'
        new_set_point['idx'] = self.idx
        new_set_point['value'] = self.w0 + d_w
        new_set_point['add'] = False
        return new_set_point

class VoltageRegulator():
    def __init__(self, **kwargs):
        self.PI_voltage = kwargs.get('pi_e', None)
        self.PI_reactive = kwargs.get('pi_q', None)
        self.b = kwargs.get('pi_q', 1)
        self.neighbors = kwargs.get('neighbors', [])

    def get_power(self, idx, system = None):
        if system is None:
            #when in colmena change this to access the data differently
            return
        
        uid = system.GENROU.idx2uid(idx)
        res = system.GENROU.p.v[uid]
        return res

    def compute_input(self, q_i, system =None):
        res = 0
        p_neighbors = [self.get_power(idx, system) for idx in self.neighbors]
        for q_j in p_neighbors:
            res = (q_j - q_i)
        
        return self.b*res
    
    def get_set_point(self, input1, input2, system):
        q_i = input2
        input2 = self.compute_input2(q_i, system)

        d_u1 = self.PI_voltage.apply(input1)
        d_u2 = self.PI_reactive.apply(input2)

        new_set_point = {}
        new_set_point['model'] = 'PVD1'
        new_set_point['param'] = 'uref0'
        new_set_point['idx'] = self.idx
        new_set_point['value'] = self.u0 + d_u1 + d_u2
        new_set_point['add'] = False

class DistributedVoltage():
    def __init__(self, **kwargs):
        self.params = {}
        self.dt = kwargs.get("dt", 0.01)
        self.Kp = kwargs.get("Kp", 0.1)  # Proportional gain
        self.Ki = kwargs.get("Ki", 0.1)  # Integral gain
        self.Lmin = kwargs.get("Lmin", -10.0)  # Minimum limit
        self.Lmax = kwargs.get("Lmax", 10.0)  # Maximum limit
        self.reference = kwargs.get("ref", 1.0)  
        self.is_delta = kwargs.get("ref", True)  
        self.model = kwargs.get("model", 'GENROU')  
        self.target_var = kwargs.get("target_var", 'paux0')  
        self.idx = kwargs.get("idx", 1)  
        
        self.mu = 0
        self.v = 0
        self.epsilon = 0
        self.epsilon_min = 0
        self.epsilon_max = 0
        self.q = 0
        self.q_min = 0.9
        self.q_max = 1.1
        #self.D = self.Df(v_0, i)
        #self.g = N*g_av(0)

    def initialise_variables(self):
        return
    
    def update_1(self):
        return