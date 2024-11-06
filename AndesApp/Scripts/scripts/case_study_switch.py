import os, sys

current_directory = os.path.dirname(os.path.abspath(__file__))
two_levels_up = os.path.dirname(os.path.dirname(current_directory))
sys.path.insert(0, two_levels_up)
import numpy as np
import pandas as pd
import matplotlib
import matplotlib.pyplot as plt
import andes_methods as ad_methods
import openpyxl
import aux_function as aux
from scipy.integrate import odeint
from scipy.optimize import root
import os, sys
#import tensorflow as tf
from andes.utils.paths import get_case, cases_root, list_cases
import andes as ad
from scipy.optimize import approx_fprime

# Now you can import your package
if __name__ == '__main__':
    _ = 0
    import andes as ad
    from andes.utils.paths import get_case, cases_root, list_cases

ad.config_logger(30)
list_cases()
base_case = False
kundur_modified_case_a = True
kundur_modified_case_b = False
print(ad.__file__)

#we first make the mecessary modifications to the xlsx document
file_path = 'kundur_lines.xlsx'
workbook = openpyxl.load_workbook(file_path)
area_sheet = workbook['Area']
bus_sheet = workbook['Bus']
area_sheet.title = 'Areac' 
old_name = 'area'
new_name = 'areac'
new_sheet_name = 'Neighbourhood'
new_sheet_name2 = 'Toggle_Line'
data = {'uid':[0,1], 'idx':[1,2], 'bus':[1, 4]}
data1 = {'uid':[0,1], 'idx':[1,2], 'bus':[1, 4]}

#we plot the results
ss = ad.load('kundur_lines.xlsx')
ss.TDS_colmena.load_plotter()

output_dir = os.path.join('plots', 'plots_csswitch')
aux.set_config(ss, setup=5)
os.makedirs(output_dir, exist_ok=True)
ss.PFlow.run()
ss.TDS_colmena.run()
ss.TDS_colmena.load_plotter()
matplotlib.use('TkAgg')


for a in ss.Line._all_vars():
    fig, ax = ss.TDS_colmena.plt.plot(getattr(ss.Line, a), a=0)
    output_path1 = os.path.join(output_dir, f'{a}.pgf')
    output_path2 = os.path.join(output_dir, f'{a}.png')
    fig.savefig(output_path1)
    fig.savefig(output_path2)
    matplotlib.pyplot.close()

for col in bus_sheet.iter_cols(1, bus_sheet.max_column, 1, 1):
    for cell in col:
        if cell.value == old_name:
            column_index = cell.column_letter
            break
        
if column_index:
    bus_sheet[f'{column_index}1'].value = new_name
    print(f'Column "{old_name}" renamed to "{new_name}"')
else:
    _=0

aux.add_sheet(workbook, new_sheet_name, data)
data2 = aux.delete_last_row(workbook, sheet_name='Line')
aux.add_sheet(workbook, new_sheet_name2, data2)
workbook.save('kundur_lines_mod2.xlsx')
ss = ad.load('kundur_lines_mod2.xlsx')

ss.PFlow.run()
aux.set_config(ss, setup=3)
ss.TDS_colmena.run()

ss.TDS_colmena.load_plotter()
matplotlib.use('TkAgg')
output_dir = os.path.join('plots', 'plots_csswitch')
os.makedirs(output_dir, exist_ok=True)

#we plot the results
for a in ss.Line._all_vars():
    fig, ax = ss.TDS_colmena.plt.plot(getattr(ss.PVD1, a), a=0)
    output_path1 = os.path.join(output_dir, f'{a}.pgf')
    output_path2 = os.path.join(output_dir, f'{a}.pnf')
    fig.savefig(output_path1)
    fig.savefig(output_path2)
    matplotlib.pyplot.close()
#plt.show()