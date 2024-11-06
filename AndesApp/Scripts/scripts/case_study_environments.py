import os, sys

current_directory = os.path.dirname(os.path.abspath(__file__))
two_levels_up = os.path.dirname(os.path.dirname(current_directory))
sys.path.insert(0, two_levels_up)
import numpy as np
import pandas as pd
import matplotlib
import matplotlib.pyplot as plt
import andes_methods as ad_methods
import openpyxl
import aux_function as aux
from scipy.integrate import odeint
from scipy.optimize import root
import os, sys
#import tensorflow as tf
from andes.utils.paths import get_case, cases_root, list_cases
import andes as ad
from scipy.optimize import approx_fprime

# Now you can import your package
if __name__ == '__main__':
    _ = 0
    import andes as ad
    from andes.utils.paths import get_case, cases_root, list_cases

ad.config_logger(30)
list_cases()
base_case = False
kundur_modified_case_a = True
kundur_modified_case_b = False
print(ad.__file__)

#we first make the mecessary modifications to the xlsx document
file_path = 'kundur_lines.xlsx'
workbook = openpyxl.load_workbook(file_path)
area_sheet = workbook['Area']
bus_sheet = workbook['Bus']
area_sheet.title = 'Area_c' 
old_name = 'area'
new_name = 'area_c'
new_sheet_name = 'neighbourhood'
data = {'uid':[0,1], 'idx':[1,2], 'bus':[1, 4]}
 
for col in bus_sheet.iter_cols(1, bus_sheet.max_column, 1, 1):
    for cell in col:
        if cell.value == old_name:
            column_index = cell.column_letter
            break
        
if column_index:
    bus_sheet[f'{column_index}1'].value = new_name
    print(f'Column "{old_name}" renamed to "{new_name}"')
else:
    _=0

aux.add_sheet(workbook, new_sheet_name, data)
workbook.save('kundur_lines_mod.xlsx')

matplotlib.use('TkAgg')
ss = ad.load('kundur_lines_mod.xlsx')
ss.PFlow.run()
ss.TDS_colmena.config.mdl = 'Toggle_Line'
ss.TDS_colmena.config.p = 'connect'
ss.TDS_colmena.config.change = 'connect'
ss.TDS_colmena.run()
#plt.show()